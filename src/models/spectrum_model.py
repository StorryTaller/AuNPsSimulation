from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice, SchemeColor
from openpyxl.drawing.line import LineProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from src.models.metrics_calc import (
    build_summary_row,
    calculate_batch_metrics,
    calculate_spectrum_metrics,
    normalize_rows,
)
from src.utils.export_helpers import format_legend_param_value, round_export_metric_row
from src.utils.runtime_paths import get_app_root


def _try_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        if isinstance(value, float) and np.isnan(value):
            return None
        return float(value)
    except Exception:
        return None


def _scale_to_nm(values: np.ndarray) -> tuple[np.ndarray, str]:
    if values.size == 0:
        return values, ""
    max_val = float(np.nanmax(values))
    if max_val < 1e-6:
        return values * 1e9, "nm"
    if max_val < 1:
        return values * 1e3, "nm"
    return values, ""


@dataclass
class SpectrumDataset:
    _METRICS_TEMPLATE_CANDIDATES = (
        get_app_root() / "res" / "templates" / "metrics_chart_template.xlsx",
        get_app_root() / "res" / "icons" / "example.xlsx",
    )

    file_path: str | None = None
    original_header: list[Any] | None = None
    wavelengths_nm: np.ndarray | None = None
    param_name: str = "Parameter"
    param_values_raw: list[Any] = field(default_factory=list)
    param_values_scaled: list[Any] | None = None
    param_unit: str = ""
    spectra_data: np.ndarray | None = None

    def load_csv_data(self, file_path: str) -> None:
        df = pd.read_csv(file_path, header=None)
        if df.shape[0] < 2 or df.shape[1] < 2:
            raise ValueError("csv 至少需要 2 行 2 列")

        self.file_path = file_path
        self.original_header = df.iloc[0].tolist()

        wl = df.iloc[0, :-1].astype(float).to_numpy(dtype=float)
        wl, _ = _scale_to_nm(wl)
        self.wavelengths_nm = wl

        self.param_name = str(df.iloc[0, -1]).strip() or "Parameter"
        self.param_values_raw = df.iloc[1:, -1].tolist()
        self.spectra_data = df.iloc[1:, :-1].astype(float).to_numpy(dtype=float)

        self._prepare_param_values()

    def _prepare_param_values(self) -> None:
        numeric = [_try_float(v) for v in self.param_values_raw]
        if not numeric or any(v is None for v in numeric):
            self.param_values_scaled = None
            self.param_unit = ""
            return

        values = np.asarray([float(v) for v in numeric], dtype=float)
        is_index = self.param_name.strip().lower() == "index"
        if is_index:
            self.param_values_scaled = values.tolist()
            self.param_unit = ""
            return

        scaled, unit = _scale_to_nm(values)
        self.param_values_scaled = scaled.tolist()
        self.param_unit = unit

    def get_param_value(self, index: int) -> Any:
        if self.param_values_scaled is not None and 0 <= index < len(self.param_values_scaled):
            return self.param_values_scaled[index]
        if 0 <= index < len(self.param_values_raw):
            return self.param_values_raw[index]
        return None

    def get_spectrum(self, index: int) -> tuple[np.ndarray, np.ndarray] | tuple[None, None]:
        if self.spectra_data is None or self.wavelengths_nm is None:
            return None, None
        if index < 0 or index >= len(self.spectra_data):
            return None, None
        return self.wavelengths_nm, self.spectra_data[index]

    def normalize_spectra(self) -> np.ndarray | None:
        if self.spectra_data is None:
            return None
        return normalize_rows(self.spectra_data)

    def save_normalized_csv(self, output_path: str | None = None) -> tuple[bool, str]:
        if self.spectra_data is None or self.wavelengths_nm is None:
            return False, "暂无光谱数据"

        normalized = self.normalize_spectra()
        if normalized is None:
            return False, "归一化失败"

        if output_path is None:
            base = Path(self.file_path or "normalized.csv")
            output_path = str(base.with_name(f"{base.stem}N.csv"))

        header = self.original_header or (self.wavelengths_nm.tolist() + [self.param_name])
        out = pd.DataFrame(np.column_stack([normalized, np.asarray(self.param_values_raw, dtype=object)]))
        result = pd.concat([pd.DataFrame([header]), out], ignore_index=True)
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        result.to_csv(output_path, index=False, header=False)
        return True, output_path

    def calculate_spectrum_metrics(self, index: int):
        if self.wavelengths_nm is None or self.spectra_data is None:
            return None
        if index < 0 or index >= len(self.spectra_data):
            return None
        return calculate_spectrum_metrics(self.wavelengths_nm, self.spectra_data[index])

    def calculate_all_metrics(self) -> dict[str, Any]:
        if self.wavelengths_nm is None or self.spectra_data is None:
            return {}
        params = self.param_values_scaled if self.param_values_scaled is not None else self.param_values_raw
        return calculate_batch_metrics(
            self.wavelengths_nm,
            self.spectra_data,
            param_values=params,
            param_name=self.param_name,
        )

    def get_batch_summary(self) -> dict[str, float]:
        return build_summary_row(self.calculate_all_metrics())

    @staticmethod
    def _write_metrics_table_to_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    @staticmethod
    def _set_axis_title_text(axis: Any, text: str) -> None:
        title = getattr(axis, "title", None)
        try:
            if title is not None and title.tx is not None and title.tx.rich is not None and title.tx.rich.p:
                first_paragraph = title.tx.rich.p[0]
                if first_paragraph.r:
                    first_paragraph.r[0].t = text
                    return
        except Exception:
            pass
        axis.title = text

    @staticmethod
    def _apply_metrics_sheet_layout(ws: Worksheet) -> None:
        ws.sheet_format.defaultRowHeight = 14.25
        ws.column_dimensions["A"].width = 7.75
        ws.column_dimensions["B"].width = 13.0
        ws.column_dimensions["C"].width = 12.75
        ws.column_dimensions["D"].width = 13.0
        ws.column_dimensions["E"].width = 13.0
        ws.column_dimensions["F"].width = 14.5
        ws.column_dimensions["G"].width = 13.25
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        ws.column_dimensions["A"].bestFit = True
        ws.column_dimensions["B"].bestFit = False
        ws.column_dimensions["C"].bestFit = True
        ws.column_dimensions["D"].bestFit = False
        ws.column_dimensions["E"].bestFit = False
        ws.column_dimensions["F"].bestFit = True
        ws.column_dimensions["G"].bestFit = True
        ws.sheet_view.tabSelected = True
        if ws.sheet_view.selection:
            ws.sheet_view.selection[0].activeCell = "G10"
            ws.sheet_view.selection[0].sqref = "G10"

    @classmethod
    def _create_metrics_workbook(cls) -> Workbook:
        for template_path in cls._METRICS_TEMPLATE_CANDIDATES:
            if not template_path.exists():
                continue
            try:
                return load_workbook(template_path)
            except Exception:
                continue
        wb = Workbook()
        wb.active.title = "Sheet1"
        return wb

    @staticmethod
    def _enforce_reference_chart_style(chart: BarChart) -> None:
        # 图表区边框：无线条
        chart_gp = chart.graphical_properties
        if chart_gp is None:
            chart_gp = GraphicalProperties()
            chart.graphical_properties = chart_gp
        if chart_gp.line is None:
            chart_gp.line = LineProperties()
        chart_gp.line.noFill = True
        chart_gp.line.prstDash = "solid"

        # 绘图区边框：无线条
        plot_gp = chart.plot_area.graphicalProperties
        if plot_gp is None:
            plot_gp = GraphicalProperties()
            chart.plot_area.graphicalProperties = plot_gp
        if plot_gp.line is None:
            plot_gp.line = LineProperties()
        plot_gp.line.noFill = True
        plot_gp.line.prstDash = "solid"

        # 主要网格线：按示例设置（实线、0.75 磅、圆角连接）
        grid = chart.y_axis.majorGridlines
        if grid is None:
            grid = ChartLines()
            chart.y_axis.majorGridlines = grid
        if grid.spPr is None:
            grid.spPr = GraphicalProperties()
        if grid.spPr.line is None:
            grid.spPr.line = LineProperties()
        line = grid.spPr.line
        line.w = 9525  # 0.75pt
        line.cap = "flat"
        line.cmpd = "sng"
        line.algn = "ctr"
        line.solidFill = ColorChoice(schemeClr=SchemeColor(val="bg1", lumMod=85000))
        line.prstDash = "solid"
        line.round = True

    def _sync_q_chart_with_sheet(self, ws: Worksheet, rows_count: int) -> None:
        if rows_count <= 0:
            return

        max_row = rows_count + 1  # header + data
        chart = ws._charts[0] if ws._charts else None
        if chart is None:
            self._append_q_chart(ws, rows_count)
            chart = ws._charts[0] if ws._charts else None
            if chart is None:
                return

        chart.title = None
        chart.style = 2
        chart.roundedCorners = True
        chart.varyColors = False
        chart.legend = None
        chart.gapWidth = 219
        chart.overlap = -27

        self._set_axis_title_text(chart.y_axis, "Q")
        self._set_axis_title_text(chart.x_axis, self.param_name)
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        if chart.x_axis.numFmt is not None:
            chart.x_axis.numFmt.formatCode = "General"
            chart.x_axis.numFmt.sourceLinked = True
        else:
            chart.x_axis.numFmt = "General"
        if chart.y_axis.numFmt is not None:
            chart.y_axis.numFmt.formatCode = "General"
            chart.y_axis.numFmt.sourceLinked = True
        else:
            chart.y_axis.numFmt = "General"
        chart.x_axis.majorTickMark = "none"
        chart.y_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.lblOffset = 100
        chart.x_axis.crosses = "autoZero"
        chart.y_axis.crosses = "autoZero"
        chart.y_axis.crossBetween = "between"
        if chart.x_axis.title is not None:
            chart.x_axis.title.overlay = False
        if chart.y_axis.title is not None:
            chart.y_axis.title.overlay = False
        self._enforce_reference_chart_style(chart)

        chart.dLbls = DataLabelList()
        chart.dLbls.showLegendKey = False
        chart.dLbls.showVal = False
        chart.dLbls.showCatName = False
        chart.dLbls.showSerName = False
        chart.dLbls.showPercent = False
        chart.dLbls.showBubbleSize = False

        val_formula = f"{ws.title}!$E$2:$E${max_row}"
        cat_formula = f"{ws.title}!$B$2:$B${max_row}"
        if chart.series:
            series = chart.series[0]
            series.tx = None
            if series.val is not None and series.val.numRef is not None:
                series.val.numRef.f = val_formula
                series.val.numRef.numCache = None
            if series.cat is not None:
                if series.cat.numRef is not None:
                    series.cat.numRef.f = cat_formula
                    series.cat.numRef.numCache = None
                elif series.cat.strRef is not None:
                    series.cat.strRef.f = cat_formula
                    series.cat.strRef.strCache = None
        else:
            values = Reference(ws, min_col=5, min_row=2, max_row=max_row)  # E: Q
            categories = Reference(ws, min_col=2, min_row=2, max_row=max_row)  # B: 参数值
            chart.add_data(values, titles_from_data=False)
            chart.set_categories(categories)

        if chart.series:
            series0 = chart.series[0]
            series0.graphicalProperties.solidFill = ColorChoice(
                schemeClr=SchemeColor(val="accent1")
            )
            if series0.graphicalProperties.line is not None:
                series0.graphicalProperties.line.noFill = True
                series0.graphicalProperties.line.prstDash = "solid"

        chart.anchor = f"A{max_row + 2}"
        chart.height = 7.5
        chart.width = 15

    def _build_metrics_rows(self) -> list[dict[str, Any]]:
        batch = self.calculate_all_metrics()
        single = batch.get("single_metrics", [])
        ris = float(batch.get("ris", 0.0) or 0.0)
        is_index = self.param_name.lower() == "index"

        rows: list[dict[str, Any]] = []
        for i, metric in enumerate(single):
            value = self.get_param_value(i)
            if metric is None:
                row: dict[str, Any] = {
                    "参数名": self.param_name,
                    "参数值": value,
                    "λ(nm)": "N/A",
                    "FWHM(nm)": "N/A",
                    "Q": "N/A",
                }
                if is_index:
                    row["RIS(nm/RIU)"] = ris
                    row["FOM(1/RIU)"] = "-"
                rows.append(row)
                continue

            fom = abs(ris) / metric.fwhm if metric.fwhm > 0 else 0.0
            row = {
                "参数名": self.param_name,
                "参数值": value,
                "λ(nm)": metric.resonance_wavelength,
                "FWHM(nm)": metric.fwhm,
                "Q": metric.q_factor,
            }
            if is_index:
                row["RIS(nm/RIU)"] = ris
                row["FOM(1/RIU)"] = fom
            rows.append(row)
        return rows

    def _append_q_chart(self, ws: Worksheet, rows_count: int) -> None:
        if rows_count <= 0:
            return

        max_row = rows_count + 1  # header + data
        chart = BarChart()
        chart.type = "col"
        chart.style = 2
        chart.title = None
        chart.y_axis.title = "Q"
        chart.x_axis.title = self.param_name
        chart.legend = None
        chart.gapWidth = 219
        chart.overlap = -27

        # Force axis/label placement so exported chart matches the Excel reference:
        # - axis numbers visible
        # - axis titles shown outside tick labels (not overlaying plot area)
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.numFmt = "General"
        chart.y_axis.numFmt = "General"
        chart.x_axis.majorTickMark = "none"
        chart.y_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"
        chart.y_axis.minorTickMark = "none"
        chart.x_axis.lblOffset = 100
        if chart.x_axis.title is not None:
            chart.x_axis.title.overlay = False
        if chart.y_axis.title is not None:
            chart.y_axis.title.overlay = False

        self._enforce_reference_chart_style(chart)

        values = Reference(ws, min_col=5, min_row=2, max_row=max_row)  # E: Q
        categories = Reference(ws, min_col=2, min_row=2, max_row=max_row)  # B: 参数值
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(categories)

        chart.height = 7.5
        chart.width = 15

        start_row = max_row + 2
        ws.add_chart(chart, f"A{start_row}")

    def export_metrics_table(
        self,
        output_path: str | None = None,
        mode: str = "multi",
        index: int | None = None,
    ) -> tuple[bool, str]:
        if self.spectra_data is None:
            return False, "暂无光谱数据"

        rows = self._build_metrics_rows()
        if not rows:
            return False, "暂无可导出的指标"

        if output_path is None:
            base = Path(self.file_path or "metrics.csv")
            output_path = str(base.with_suffix(".xlsx"))

        target = Path(output_path)
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")

        target.parent.mkdir(parents=True, exist_ok=True)
        formatted_rows = [round_export_metric_row(row, decimals=3) for row in rows]
        df = pd.DataFrame(formatted_rows)

        wb = self._create_metrics_workbook()
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        ws.title = "Sheet1"
        self._write_metrics_table_to_sheet(ws, df)
        self._apply_metrics_sheet_layout(ws)
        self._sync_q_chart_with_sheet(ws, len(df))
        wb.save(target)
        return True, str(target)


@dataclass
class MultiDimSpectrumDataset:
    file_path: str | None = None
    wavelengths_nm: np.ndarray | None = None
    spectra_data_full: np.ndarray | None = None
    params_df_full: pd.DataFrame | None = None
    param_names: list[str] = field(default_factory=list)
    param_unit_map: dict[str, str] = field(default_factory=dict)

    def load_csv_multidim(self, file_path: str) -> None:
        df = pd.read_csv(file_path, header=None)
        if df.shape[0] < 2:
            raise ValueError("csv 至少需要 2 行")

        self.file_path = file_path
        header = df.iloc[0].tolist()

        last_wl_idx = -1
        for i in range(len(header) - 1, -1, -1):
            if _try_float(header[i]) is not None:
                last_wl_idx = i
                break
        if last_wl_idx < 0:
            raise ValueError("未找到波长列")

        wl = np.asarray([float(header[i]) for i in range(last_wl_idx + 1)], dtype=float)
        wl, _ = _scale_to_nm(wl)
        self.wavelengths_nm = wl

        data_rows = df.iloc[1:, :]
        self.spectra_data_full = data_rows.iloc[:, : last_wl_idx + 1].astype(float).to_numpy(dtype=float)

        raw_names = [
            str(v).strip() if str(v).strip() not in {"", "nan", "None"} else f"param_{i+1}"
            for i, v in enumerate(header[last_wl_idx + 1 :])
        ]
        if not raw_names:
            raise ValueError("未找到参数列")

        uniq_names: list[str] = []
        seen: dict[str, int] = {}
        for name in raw_names:
            seen[name] = seen.get(name, 0) + 1
            uniq_names.append(name if seen[name] == 1 else f"{name}_{seen[name]}")
        self.param_names = uniq_names

        params: dict[str, list[Any]] = {}
        for i, name in enumerate(self.param_names):
            col_idx = last_wl_idx + 1 + i
            params[name] = data_rows.iloc[:, col_idx].tolist()
        self.params_df_full = pd.DataFrame(params)

        self._infer_param_units()

    def _infer_param_units(self) -> None:
        self.param_unit_map = {}
        if self.params_df_full is None:
            return

        for name in self.param_names:
            values = self.params_df_full[name].tolist()
            numeric = [_try_float(v) for v in values]
            if not numeric or any(v is None for v in numeric):
                self.param_unit_map[name] = "raw"
                continue

            if name.strip().lower() == "index":
                self.param_unit_map[name] = "index"
                continue

            arr = np.asarray([float(v) for v in numeric], dtype=float)
            max_val = float(np.nanmax(arr))
            if max_val < 1e-6:
                self.param_unit_map[name] = "m"
            elif max_val < 1:
                self.param_unit_map[name] = "um"
            else:
                self.param_unit_map[name] = "raw"

    def get_param_unique_values(self, param_name: str) -> list[Any]:
        if self.params_df_full is None or param_name not in self.param_names:
            return []
        values = [v for v in self.params_df_full[param_name].tolist() if not pd.isna(v)]
        unique = list(dict.fromkeys(values))
        numeric = [_try_float(v) for v in unique]
        if numeric and all(v is not None for v in numeric):
            unique = [
                pair[1]
                for pair in sorted(zip([float(v) for v in numeric], unique), key=lambda item: item[0])
            ]
        return unique

    def convert_param_value(self, param_name: str, raw_value: Any) -> Any:
        f = _try_float(raw_value)
        if f is None:
            return raw_value
        unit = self.param_unit_map.get(param_name, "raw")
        if unit == "index":
            return float(f)
        if unit == "m":
            return float(f) * 1e9
        if unit == "um":
            return float(f) * 1e3
        return float(f)

    def format_param_value(
        self,
        param_name: str,
        raw_value: Any,
        decimals: int = 4,
        compact: bool = False,
    ) -> str:
        value = self.convert_param_value(param_name, raw_value)
        if isinstance(value, float):
            if compact:
                text = format_legend_param_value(value, decimals=decimals)
            else:
                text = f"{value:.{int(decimals)}f}"
            if self.param_unit_map.get(param_name, "raw") in {"m", "um"}:
                return f"{text} nm"
            return text
        return str(value)

    def _match_mask(self, selection: dict[str, Any]) -> np.ndarray:
        if self.params_df_full is None:
            raise ValueError("参数数据未加载")
        mask = np.ones(len(self.params_df_full), dtype=bool)
        for key, value in selection.items():
            if key not in self.param_names or value is None:
                continue
            mask &= self.params_df_full[key].values == value
        return mask

    def get_spectrum_by_selection(self, selection: dict[str, Any]) -> tuple[np.ndarray, np.ndarray, int]:
        if self.spectra_data_full is None or self.wavelengths_nm is None:
            raise ValueError("光谱数据未加载")
        mask = self._match_mask(selection)
        idx = np.where(mask)[0]
        if len(idx) == 0:
            raise ValueError("未找到匹配的参数组合")
        row = int(idx[0])
        return self.wavelengths_nm, self.spectra_data_full[row], row

    def get_sweep(
        self,
        varying_param: str,
        fixed_params: dict[str, Any],
        selected_values: list[Any] | None = None,
    ) -> tuple[np.ndarray, np.ndarray, list[Any]]:
        if self.params_df_full is None or self.spectra_data_full is None or self.wavelengths_nm is None:
            raise ValueError("多维数据未加载")
        if varying_param not in self.param_names:
            raise ValueError(f"参数不存在: {varying_param}")

        selection = {k: v for k, v in fixed_params.items() if k != varying_param}
        mask = self._match_mask(selection)

        df = self.params_df_full.loc[mask, :].reset_index(drop=True)
        spectra = self.spectra_data_full[mask]
        if len(df) == 0:
            raise ValueError("固定参数条件下没有匹配数据")

        varying_values = df[varying_param].tolist()
        unique = list(dict.fromkeys([v for v in varying_values if not pd.isna(v)]))
        numeric = [_try_float(v) for v in unique]
        if numeric and all(v is not None for v in numeric):
            unique = [
                pair[1]
                for pair in sorted(zip([float(v) for v in numeric], unique), key=lambda item: item[0])
            ]

        if selected_values is not None:
            picked = [v for v in selected_values if not pd.isna(v)]
            if not picked:
                raise ValueError("请至少选择一个参数值")
            unique = [value for value in unique if any(value == target for target in picked)]
            if not unique:
                raise ValueError("所选参数值在当前固定条件下无匹配数据")

        picked_spectra: list[np.ndarray] = []
        picked_values: list[Any] = []
        for value in unique:
            idx = next((i for i, v in enumerate(varying_values) if v == value), None)
            if idx is None:
                continue
            picked_spectra.append(spectra[idx])
            picked_values.append(value)

        if not picked_spectra:
            raise ValueError("未找到可绘制的光谱")

        return self.wavelengths_nm, np.asarray(picked_spectra, dtype=float), picked_values
